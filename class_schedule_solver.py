from operator import attrgetter     # Sorting with custom objects (class objects)
from openpyxl import load_workbook  # Output schedules to an Excel sheet
from itertools import permutations


class Schedules:
    """
    This class, "Schedules", is a class that creates working schedules for university classes that it is passed.

    Attributes:
        term (string): This is the specific term that the schedule is being made for.
        open_courses (list of departments): The list of departments, i.e. each department has several courses and those courses have several classes.
        workbook_for_classes (openpyxl workbook): This is a workbook that holds the schedule for the classes.
        workbook_for_finals (openpyxl workbook): This is a workdbook that holds the schedule for the finals for the classes.
        class_list_for_classes (openpyxl sheet): This is a sheet that is written to when creating the class schedule.
        class_list_for_finals (openpyxl sheet): This is a sheet that is written to when creating the finals schedule for the classes.
        days (dict): This is a dictionary that is used for quick access to full-day-names given their short abbreviations.

    """

    def __init__(self, term, open_courses):
        """
        The constructor for the Schedules class.

        :param term: A string containing the term of the classes that were scraped (e.g. Fall 2018).
        :param open_courses: A list of departments (all of the classes that were scraped).
        :return: Nothing.
        """

        self.term = term
        self.open_courses = open_courses
        self.workbook_for_classes = load_workbook('Schedule_Template.xlsx')
        self.workbook_for_finals = load_workbook('Schedule_Template.xlsx')
        self.class_list_for_classes = self.workbook_for_classes['Class List']
        self.class_list_for_finals = self.workbook_for_finals['Class List']
        self.days = {
            'M': 'MONDAY',
            'Tu': 'TUESDAY',
            'W': 'WEDNESDAY',
            'Th': 'THURSDAY',
            'F': 'FRIDAY',
            'Mon': 'MONDAY',
            'Tue': 'TUESDAY',
            'Wed': 'WEDNESDAY',
            'Thu': 'THURSDAY',
            'Fri': 'FRIDAY'
        }

    def make_a_schedule(self):
        """
        This function creates a schedule that prioritizes early classes and also writes this schedule to an Excel class schedule template.

        :return: None
        """

        almost_sorted_courses = []
        for department in self.open_courses:
            for course in department:
                almost_sorted_courses.append(course)
        sorted_courses = sorted(almost_sorted_courses, key=attrgetter('number_of_choices'))          # Sort the courses based on how many choices they have for classes (i.e. more choices are less prioritized)

        some_what_sorted_classes = []
        number_of_classes_required_total = 0
        for course in sorted_courses:
            some_what_sorted_classes += course.get_all_classes()
            number_of_classes_required_total += course.get_number_of_required_classes()

        possible_schedules = Schedules.find_schedules(some_what_sorted_classes, number_of_classes_required_total)
        filtered_schedules = Schedules.filter_schedules(possible_schedules, sorted_courses, number_of_classes_required_total)

    @staticmethod
    def filter_schedules(possible_schedules, sorted_courses, number_of_classes_required_total):
        filtered_schedules = []

        for schedule in possible_schedules.copy():
            if len(schedule) < number_of_classes_required_total:
                possible_schedules.remove(schedule)

        for schedule in possible_schedules:
            current_filtered_schedule = []
            for course in sorted_courses:
                course_name = course.get_name_of_course()
                lab_fulfillment = False
                discussion_fulfillment = False
                lecture_fulfillment = False
                for _class in schedule:
                    if course_name == _class.name_of_course:
                        class_type = _class.get_type()
                        if course.has_lab() and class_type == 'Lab':
                            if not lab_fulfillment:
                                current_filtered_schedule.append(_class)
                                lab_fulfillment = True
                        elif course.has_discussion() and class_type == 'Dis':
                            if not discussion_fulfillment:
                                current_filtered_schedule.append(_class)
                                discussion_fulfillment = True
                        elif class_type == 'Lec' and not lecture_fulfillment:
                            current_filtered_schedule.append(_class)
                            lecture_fulfillment = True
            if len(current_filtered_schedule) == number_of_classes_required_total:
                filtered_schedules.append(current_filtered_schedule)
        return filtered_schedules

    @staticmethod
    def find_schedules(class_list, number_of_classes_required_total):
        possible_schedules = []
        for class_ in class_list:
            temp_class = class_
            half_optimized_list = Schedules.remove_same_type(temp_class, class_list.copy())
            optimized_list = Schedules.remove_conflicts(temp_class, half_optimized_list.copy())
            for x in list(permutations(optimized_list, number_of_classes_required_total)):
                Schedules.filter_schedules(possible_schedules, x, number_of_classes_required_total)
                if len(possible_schedules) > 0:
                    break
        return possible_schedules

    @staticmethod
    def find_schedules_helper(current_class_list, current_schedule, collection_of_schedules):
        current_schedule.append(current_class_list[0])
        temp = current_class_list[0]
        new_class_list = Schedules.remove_conflicts(temp, current_class_list.copy())
        Schedules.remove_same_type(temp, new_class_list)
        current_class_list.remove(temp)

        if len(new_class_list) == 0:
            collection_of_schedules.append(current_schedule)
        else:
            Schedules.find_schedules_helper(new_class_list, current_schedule, collection_of_schedules)

    @staticmethod
    def remove_conflicts(chosen_class, class_list):
        #class_list.remove(chosen_class)             # It is already chosen so remove it from the list

        for _class in class_list:
            if Schedules.is_between(chosen_class, _class) and not (chosen_class.code == _class.code):
                class_list.remove(_class)
        return class_list

    @staticmethod
    def remove_same_type(chosen_class, class_list):
        class_type = chosen_class.get_type()
        name_of_course = chosen_class.get_name_of_course()
        for _class in class_list:
            if (_class.get_type() == class_type) and (_class.get_name_of_course() == name_of_course):
                if chosen_class.code == _class.code:
                    pass
                else:
                    class_list.remove(_class)
        return class_list

    @staticmethod
    def is_between(class_a, class_b):
        """
        A function to check if "class b's" time overlaps "class_a's"

        :param class_a: A class object
        :param class_b: A class object
        :return: True if it does overlap, False if the classes do not overlap
        """

        time_start_a, time_end_a = class_a.get_time()
        time_start_b, time_end_b = class_b.get_time()

        if ((time_start_a < time_start_b) and (time_end_a < time_start_b)) and (class_a.days == class_b.days):
            return False
        elif ((time_start_b < time_start_a) and (time_end_b < time_start_a)) and (class_a.days == class_b.days):
            return False
        else:
            return True

    def __del__(self):
        """
        In the destructor for the Schedules class, the workbooks (Excel sheets) are saved as the file names "Class Schedule - 'TERM'.xlsx" and "Finals Schedule - 'TERM'.xlsx".

        :return: None
        """

        self.workbook_for_classes.save('Class Schedule - ' + self.term + ".xlsx")
        self.workbook_for_finals.save('Finals Schedule - ' + self.term + ".xlsx")
