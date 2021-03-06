from openpyxl import load_workbook  # Output schedules to an Excel sheet


def change_time_to_normal(class_list):
    """
    A function to change all the time to standard time as opposed to military time.

    :param class_list: A list of Class objects.
    :return:
    """
    for _class in class_list:
        _class.update_time_to_normal_format()


def export_text(term, working_schedule):
    """
    This function will export classes to a text file so that users can read it.

    :param term: The term of which the classes were scraped from.
    :param working_schedule: Working schedules that have non-conflicting classes.
    :return: Nothing
    """
    file = open('class_schedule.txt', 'w')
    counter = 1
    file.write(term + '\n')
    for schedule in working_schedule:
        file.write('Potential Schedule #' + str(counter) + '\n')
        class_list_normal_clock = schedule.get_class_list()
        change_time_to_normal(class_list_normal_clock)
        for _class in class_list_normal_clock:
            file.write(str(_class.code) + ', '
                       + _class.name_of_course + '-'
                       + _class.type_of_class + ', '
                       + _class.instructor + ', '
                       + _class.days + ' '
                       + _class.normal_time_start + ' - '
                       + _class.normal_time_end + ' \n')
        counter += 1
        file.write('\n')
    file.close()


def export_excel(term, working_schedule):
    """
    IN PROGRESS
    :param term: A string representing the term of the web-scraped classes.
    :param working_schedule: A Schedule object.
    :return: Nothing
    """

    workbook_for_classes = load_workbook('../Schedule_Template.xlsx')
    # workbook_for_finals = load_workbook('Schedule_Template.xlsx')
    class_list_sheet_for_classes = workbook_for_classes['Class List']
    # class_list_sheet_for_finals = workbook_for_finals['Class List']
    days = {
        '': 'TBA',
        'M': 'MONDAY',
        'Tu': 'TUESDAY',
        'W': 'WEDNESDAY',
        'Th': 'THURSDAY',
        'F': 'FRIDAY',
        'Mon': 'MONDAY',
        'Tue': 'TUESDAY',
        'Wed': 'WEDNESDAY',
        'Thu': 'THURSDAY',
        'Fri': 'FRIDAY'
    }
    class_list_normal_clock = working_schedule.get_class_list()
    change_time_to_normal(class_list_normal_clock)

    current_row = 3
    starting_column = current_column = 2
    last_column = 7
    row_for_tba = 27
    used_tba = False

    for _class in class_list_normal_clock:
        days_for_class = _class.get_days_as_list()
        name_and_type = (_class.name_of_course + '-' + _class.type_of_class)
        code = _class.code
        place = _class.place
        start = _class.normal_time_start
        end = _class.normal_time_end

        for day in days_for_class:
            while current_column <= last_column:
                if start == 'TBA':
                    used_tba = True
                    current_cell = class_list_sheet_for_classes.cell(row=row_for_tba, column=current_column)
                else:
                    current_cell = class_list_sheet_for_classes.cell(row=current_row, column=current_column)
                if current_column == 2:  # Name
                    current_cell.value = code
                elif current_column == 3:  # Class code
                    current_cell.value = name_and_type
                elif current_column == 4:  # Day
                    current_cell.value = days[day]
                elif current_column == 5:  # Location
                    current_cell.value = place
                elif current_column == 6:
                    current_cell.value = start
                else:
                    current_cell.value = end
                current_column += 1
            current_column = starting_column
            if not used_tba:
                current_row += 1
            used_tba = False

    workbook_for_classes.save('../Class Schedule - ' + term + ".xlsx")
    # workbook_for_finals.save('Finals Schedule - ' + term + ".xlsx")
